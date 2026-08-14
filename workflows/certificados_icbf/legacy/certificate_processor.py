from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from math import isfinite
from numbers import Real
import os
import re
import unicodedata
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import from_excel
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader


OUTPUT_COLUMNS = [
    "N°",
    "PRIMER NOMBRE",
    "SEGUNDO NOMBRE",
    "PRIMER APELLIDO",
    "SEGUNDO APELLIDO",
    "DOCUMENTO",
    "FECHA DE NACIMIENTO",
    "UNIDADES",
]

EDITABLE_COLUMNS = [
    "INCLUIR",
    "PRIMER NOMBRE",
    "SEGUNDO NOMBRE",
    "PRIMER APELLIDO",
    "SEGUNDO APELLIDO",
    "DOCUMENTO",
    "FECHA DE NACIMIENTO",
    "UNIDADES",
    "_FILA_ORIGEN",
]

COLUMN_ALIASES = {
    "PRIMER NOMBRE": ["PRIMER NOMBRE"],
    "SEGUNDO NOMBRE": ["SEGUNDO NOMBRE"],
    "PRIMER APELLIDO": ["PRIMER APELLIDO"],
    "SEGUNDO APELLIDO": ["SEGUNDO APELLIDO"],
    "DOCUMENTO": [
        "NUMERO DE IDENTIFICACION",
        "NUMERO DOCUMENTO",
        "DOCUMENTO",
        "NRO DOCUMENTO",
    ],
    "FECHA DE NACIMIENTO": ["FECHA DE NACIMIENTO", "FECHA NACIMIENTO"],
    "UNIDADES": ["UNIDADES", "UNIDAD"],
    "TIPO DE NOVEDAD": ["TIPO DE NOVEDAD", "NOVEDAD"],
}

OPTIONAL_NA_FIELDS = {"SEGUNDO NOMBRE", "SEGUNDO APELLIDO"}
REQUIRED_FIELDS = [
    "PRIMER NOMBRE",
    "DOCUMENTO",
    "FECHA DE NACIMIENTO",
    "UNIDADES",
]
MISSING_TOKENS = {"", "NA", "N/A", "N.A.", "NONE", "NAN", "NULL"}
DEFAULT_ROWS_PER_PAGE = 25


class InputFormatError(ValueError):
    pass


def _key(value: object) -> str:
    """Normalize text for comparison and sorting.

    Removes accents, converts to uppercase, collapses whitespace, and returns a
    canonical key string.
    """
    text = "" if value is None else str(value)
    text = " ".join(text.replace("\n", " ").split()).upper()
    return "".join(
        char for char in unicodedata.normalize("NFD", text)
        if unicodedata.category(char) != "Mn"
    )


def _text(value: object) -> str:
    """Convert a value into a safe single-line string.

    Returns an empty string for missing or NaN values, and collapses extra
    whitespace for all other values.
    """
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return ""
    return " ".join(str(value).strip().split())


def _document(value: object) -> str:
    """Preserve textual identifiers and normalize only numeric Excel values."""
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return ""
    if isinstance(value, Real) and not isinstance(value, bool):
        numeric = float(value)
        if isfinite(numeric) and numeric.is_integer():
            text = str(int(numeric))
            return text.zfill(10) if len(text) <= 10 else text
    return str(value).strip()


def _document_from_excel(value: object) -> str:
    """Apply the historical zero-padding rule at the Excel ingestion boundary."""
    text = _document(value)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    if text.isdigit() and len(text) <= 10:
        return text.zfill(10)
    return text


def _sanitize_filename(name: object) -> str:
    """Prepare a safe filename segment for ZIP entries.

    Replaces unsupported characters with underscores and ensures at least one
    fallback value is returned.
    """
    text = "" if name is None else str(name)
    normalized = "".join(
        char if char.isalnum() or char in "-_" else "_"
        for char in text
    )
    return normalized.strip("_")[:128] or "documento"


def is_valid_document(value: object) -> bool:
    """Return True if the normalized document string is exactly 10 digits."""
    return bool(re.fullmatch(r"\d{10}", _document(value)))


def is_standard_document(value: object) -> bool:
    """Return True only for the standard ten-digit document format."""
    return is_valid_document(value)


def is_missing(value: object) -> bool:
    """Detect whether a value should be treated as missing or NA."""
    return _key(value) in MISSING_TOKENS


def _find_header_row(raw: pd.DataFrame) -> int:
    """Find the header row index in a raw Excel sheet.

    The file may contain introductory rows above the real headers. This helper
    scans the first 20 rows looking for a combination of 'PRIMER NOMBRE' and
    either 'NUMERO DE IDENTIFICACION' or 'DOCUMENTO'.
    """
    for row_index in range(min(len(raw), 20)):
        keys = {_key(value) for value in raw.iloc[row_index].tolist()}
        if "PRIMER NOMBRE" in keys and (
            "NUMERO DE IDENTIFICACION" in keys or "DOCUMENTO" in keys
        ):
            return row_index
    raise InputFormatError(
        "No encontré la fila de encabezados. Debe incluir Primer Nombre y "
        "Número de Identificación/Documento."
    )


def _resolve_columns(columns: list[object]) -> dict[str, object]:
    """Resolve and normalize Excel column names to the internal schema.

    Supports multiple alias names for the same logical column. Raises
    InputFormatError when required columns are missing.
    """
    normalized = {_key(column): column for column in columns}
    resolved: dict[str, object] = {}
    for target, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                resolved[target] = normalized[alias]
                break
            prefixed = next(
                (original for key, original in normalized.items() if key.startswith(alias)),
                None,
            )
            if prefixed is not None:
                resolved[target] = prefixed
                break
    required_source = {
        "PRIMER NOMBRE",
        "SEGUNDO NOMBRE",
        "PRIMER APELLIDO",
        "SEGUNDO APELLIDO",
        "DOCUMENTO",
        "FECHA DE NACIMIENTO",
        "UNIDADES",
    }
    missing = sorted(required_source - resolved.keys())
    if missing:
        raise InputFormatError(
            "Faltan estas columnas en el archivo: " + ", ".join(missing)
        )
    return resolved


def read_and_clean_excel(file_or_buffer: object) -> tuple[pd.DataFrame, dict[str, int]]:
    """Read an Excel file and normalize it into the application's working schema.

    The function detects the correct header row, validates required columns,
    filters only ingreso records, normalizes documents and dates, and prepares
    editable fields for the Streamlit UI.

    Returns a cleaned DataFrame and processing statistics.
    """
    try:
        raw = pd.read_excel(file_or_buffer, header=None, dtype=object)
    except Exception as error:  # pragma: no cover - dependemos del flujo de entrada real
        raise InputFormatError("No se pudo leer el archivo Excel. Verifica que sea un archivo válido y que no esté dañado.") from error

    try:
        header_row = _find_header_row(raw)
    except InputFormatError as error:
        raise InputFormatError(
            "No se pudo detectar la estructura del archivo. Asegúrate de usar un Excel con encabezados válidos."
        ) from error

    try:
        data = pd.read_excel(file_or_buffer, header=header_row, dtype=object)
    except Exception as error:  # pragma: no cover - dependemos del flujo de entrada real
        raise InputFormatError("No se pudo leer el archivo Excel. Verifica que sea un archivo válido y que no esté dañado.") from error

    data = data.dropna(how="all").copy()
    if data.empty:
        raise InputFormatError("El archivo Excel no contiene filas de datos para procesar.")

    try:
        resolved = _resolve_columns(list(data.columns))
    except InputFormatError as error:
        raise InputFormatError(
            "El archivo no tiene las columnas requeridas para continuar. Revisa el formato del Excel."
        ) from error

    total_received = len(data)
    if "TIPO DE NOVEDAD" in resolved:
        novelty = data[resolved["TIPO DE NOVEDAD"]].map(_key)
        keep = novelty.isin({"IN", "INGRESO"})
        data = data.loc[keep].copy()

    cleaned = pd.DataFrame(index=data.index)
    cleaned["INCLUIR"] = True
    for target in EDITABLE_COLUMNS[1:-1]:
        source = resolved[target]
        if target == "DOCUMENTO":
            cleaned[target] = data[source].map(_document_from_excel)
        elif target == "FECHA DE NACIMIENTO":
            cleaned[target] = data[source].map(format_date)
        else:
            cleaned[target] = data[source].map(_text)

    for field in OPTIONAL_NA_FIELDS:
        cleaned[field] = cleaned[field].map(lambda value: "NA" if is_missing(value) else value)

    cleaned["_FILA_ORIGEN"] = data.index.to_series().map(lambda value: int(value) + header_row + 2)
    cleaned = cleaned[EDITABLE_COLUMNS].reset_index(drop=True)

    stats = {
        "recibidos": total_received,
        "ingresos": len(cleaned),
        "excluidos": total_received - len(cleaned),
    }
    return cleaned, stats


def format_date(value: object) -> str:
    """Normalize date-like values to the expected dd/mm/YYYY string format.

    Accepts datetime objects, pandas timestamps, and parsable text. Returns an empty
    string for missing values.
    """
    if value is None or (not isinstance(value, str) and pd.isna(value)):
        return ""
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, Real) and not isinstance(value, bool):
        serial = float(value)
        if isfinite(serial) and 1 <= serial <= 2_958_465:
            try:
                return from_excel(serial).strftime("%d/%m/%Y")
            except (OverflowError, ValueError):
                pass
    text = _text(value)
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.notna(parsed):
        return parsed.strftime("%d/%m/%Y")
    return text


def normalize_edited_records(records: pd.DataFrame) -> pd.DataFrame:
    """Clean a DataFrame edited through the Streamlit data editor.

    Ensures all editable columns exist, normalizes documents and dates, and
    converts missing fields to a standardized NA marker when appropriate.
    """
    result = records.copy()
    for column in EDITABLE_COLUMNS:
        if column not in result:
            result[column] = True if column == "INCLUIR" else ""
    result["INCLUIR"] = result["INCLUIR"].fillna(False).astype(bool)
    for column in EDITABLE_COLUMNS[1:-1]:
        if column == "DOCUMENTO":
            result[column] = result[column].map(_document)
        elif column == "FECHA DE NACIMIENTO":
            result[column] = result[column].map(format_date)
        else:
            result[column] = result[column].map(_text)
    for field in OPTIONAL_NA_FIELDS:
        result[field] = result[field].map(lambda value: "NA" if is_missing(value) else value)
    return result[EDITABLE_COLUMNS].reset_index(drop=True)


def validate_records(
    records: pd.DataFrame,
    authorized_document_exceptions: set[int] | frozenset[int] = frozenset(),
) -> dict[str, pd.DataFrame | bool]:
    """Validate the current records and classify all blocking issues.

    Returns a dictionary containing:
    - active: rows marked for inclusion
    - missing_document: rows with empty or NA documents
    - invalid_document: rows with document values that do not match the required 10-digit format
    - duplicates: rows sharing the same valid document value
    - missing_report: rows with missing required fields for PDF generation
    - blocking: boolean indicating if PDF generation must be blocked
    """
    active = normalize_edited_records(records)
    active = active.loc[active["INCLUIR"]].copy()
    missing_document = active.loc[active["DOCUMENTO"].map(is_missing)].copy()
    nonstandard_document = active.loc[
        ~active["DOCUMENTO"].map(is_missing)
        & ~active["DOCUMENTO"].map(is_valid_document)
    ].copy()
    invalid_document = nonstandard_document.loc[
        ~nonstandard_document["_FILA_ORIGEN"].isin(authorized_document_exceptions)
    ].copy()

    valid_docs = active.loc[active["DOCUMENTO"].map(is_valid_document), "DOCUMENTO"]
    duplicated_values = set(valid_docs[valid_docs.duplicated(keep=False)].tolist())
    duplicates = active.loc[active["DOCUMENTO"].isin(duplicated_values)].copy()

    missing_rows: list[dict[str, object]] = []
    for _, row in active.iterrows():
        missing = [field for field in REQUIRED_FIELDS if is_missing(row[field])]
        if is_missing(row["PRIMER APELLIDO"]) and is_missing(row["SEGUNDO APELLIDO"]):
            missing.append("PRIMER APELLIDO o SEGUNDO APELLIDO")
        if missing:
            row_data = row.to_dict()
            row_data["CAMPOS OBLIGATORIOS FALTANTES"] = ", ".join(missing)
            missing_rows.append(row_data)
    missing_report = pd.DataFrame(
        missing_rows,
        columns=[*active.columns.tolist(), "CAMPOS OBLIGATORIOS FALTANTES"],
    )
    blocking = (
        not missing_report.empty
        or not invalid_document.empty
        or not duplicates.empty
    )
    return {
        "active": active,
        "missing_document": missing_document,
        "invalid_document": invalid_document,
        "nonstandard_document": nonstandard_document,
        "duplicates": duplicates,
        "missing_report": missing_report,
        "blocking": blocking,
    }


def final_records(
    records: pd.DataFrame,
    authorized_document_exceptions: set[int] | frozenset[int] = frozenset(),
) -> pd.DataFrame:
    """Return the final sorted and numbered records that will be printed to PDF.

    This function ensures the current records are fully valid for PDF generation,
    sorts them first by unit and then by name, and adds a sequential "N°" column.
    """
    validation = validate_records(records, authorized_document_exceptions)
    if validation["blocking"]:
        raise ValueError(
            "Todavía existen campos obligatorios faltantes, documentos inválidos "
            "o documentos duplicados."
        )
    active = validation["active"].copy()
    active["_ORDEN_UNIDAD"] = active["UNIDADES"].map(_key)
    active["_ORDEN_NOMBRE"] = (
        active["PRIMER APELLIDO"].map(_key) + " " + active["PRIMER NOMBRE"].map(_key)
    )
    active = active.sort_values(
        ["_ORDEN_UNIDAD", "_ORDEN_NOMBRE", "DOCUMENTO"], kind="stable"
    ).reset_index(drop=True)
    output = active[[column for column in OUTPUT_COLUMNS if column != "N°"]].copy()
    output.insert(0, "N°", range(1, len(output) + 1))
    return output


def dataframe_to_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Render one or more DataFrame sheets into an in-memory Excel file.

    Builds a polished workbook with bold headers, autofilter, frozen top row, and
    adaptive column widths suitable for downloads.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)
    header_fill = PatternFill("solid", fgColor="99CCFF")
    for sheet_name, frame in sheets.items():
        sheet = workbook.create_sheet(title=sheet_name[:31])
        for column_index, column in enumerate(frame.columns, start=1):
            cell = sheet.cell(1, column_index, column)
            cell.font = Font(name="Arial", size=11, bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
            for column_index, value in enumerate(row, start=1):
                cell = sheet.cell(row_index, column_index, "" if pd.isna(value) else value)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _register_font() -> tuple[str, str]:
    """Register a PDF font pair for ReportLab.

    Attempts to use system Arial fonts on Windows or common Linux fallback fonts.
    If no TTF fonts are found, falls back to built-in Helvetica.
    """
    candidates: list[tuple[Path, Path]] = []
    windows_dir = os.environ.get("WINDIR")
    if windows_dir:
        fonts_dir = Path(windows_dir) / "Fonts"
        candidates.append((fonts_dir / "arial.ttf", fonts_dir / "arialbd.ttf"))
    candidates.extend([
        ("/usr/share/fonts/truetype/msttcorefonts/Arial.ttf", "/usr/share/fonts/truetype/msttcorefonts/Arial_Bold.ttf"),
        ("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf", "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
    ])
    for regular, bold in candidates:
        regular_path, bold_path = Path(regular), Path(bold)
        if regular_path.exists() and bold_path.exists():
            pdfmetrics.registerFont(TTFont("CertificateArial", str(regular_path)))
            pdfmetrics.registerFont(TTFont("CertificateArialBold", str(bold_path)))
            return "CertificateArial", "CertificateArialBold"
    return "Helvetica", "Helvetica-Bold"


def _paginate(frame: pd.DataFrame, rows_per_page: int = DEFAULT_ROWS_PER_PAGE) -> list[list[dict[str, object]]]:
    """Split records into pages while preserving UNIDADES groups.

    Ensures that unit groups are not split across pages unless they exceed the page
    capacity, in which case the group is split cleanly.
    """
    pages: list[list[dict[str, object]]] = []
    current: list[dict[str, object]] = []
    groups: list[list[dict[str, object]]] = []
    for _, group in frame.groupby("UNIDADES", sort=False, dropna=False):
        groups.append(group.to_dict("records"))

    for group in groups:
        if len(group) <= rows_per_page:
            remaining = rows_per_page - len(current)
            if current and len(group) > remaining:
                pages.append(current)
                current = []
            current.extend(group)
            continue

        if current:
            pages.append(current)
            current = []
        for start in range(0, len(group), rows_per_page):
            chunk = group[start:start + rows_per_page]
            if len(chunk) == rows_per_page:
                pages.append(chunk)
            else:
                current = chunk
    if current:
        pages.append(current)
    return pages or [[]]


def _fit_text(pdf: canvas.Canvas, text: object, max_width: float, font: str, size: float) -> float:
    """Choose the largest font size that fits text inside the available width."""
    value = _text(text)
    candidate = size
    while candidate > 5.5 and pdfmetrics.stringWidth(value, font, candidate) > max_width:
        candidate -= 0.25
    return candidate


def _build_pdf_bytes(
    frame: pd.DataFrame,
    rows_per_page: int = DEFAULT_ROWS_PER_PAGE,
    logo_path: str | Path | None = None,
) -> bytes:
    """Create a PDF from a validated set of records.

    Renders one or more pages into a BytesIO buffer, with a table header and
    dynamic row height based on available page space.
    """
    try:
        pages = _paginate(frame, rows_per_page=rows_per_page)
        buffer = BytesIO()
        page_width, page_height = A4
        pdf = canvas.Canvas(buffer, pagesize=(page_width, page_height))
        regular_font, bold_font = _register_font()

        left = 25
        logo = None
        logo_width = 0.0
        logo_height = 0.0
        if logo_path:
            logo_file = Path(logo_path)
            if logo_file.is_file():
                logo = ImageReader(str(logo_file))
                source_width, source_height = logo.getSize()
                max_width, max_height = 120.0, 48.0
                scale = min(max_width / source_width, max_height / source_height)
                logo_width = source_width * scale
                logo_height = source_height * scale

        top = page_height - (logo_height + 45 if logo else 32)
        header_height = 34
        footer_margin = 20
        widths = [22, 65, 68, 65, 68, 74, 74, 109]
        headers = OUTPUT_COLUMNS

        for page_number, page_rows in enumerate(pages, start=1):
            if logo:
                pdf.drawImage(
                    logo,
                    left,
                    page_height - 20 - logo_height,
                    width=logo_width,
                    height=logo_height,
                    preserveAspectRatio=True,
                    mask="auto",
                )

            rows_count = len(page_rows)
            available_height = top - footer_margin - header_height
            row_height = min(19.0, max(7.0, available_height / max(rows_count, 1)))

            y = top
            pdf.setFillColor(colors.HexColor("#99CCFF"))
            pdf.rect(left, y - header_height, sum(widths), header_height, fill=1, stroke=0)

            x = left
            pdf.setStrokeColor(colors.black)
            pdf.setLineWidth(0.65)
            for width, header in zip(widths, headers):
                pdf.rect(x, y - header_height, width, header_height, fill=0, stroke=1)
                pdf.setFillColor(colors.black)
                pdf.setFont(bold_font, 8)
                if header == "FECHA DE NACIMIENTO":
                    pdf.drawCentredString(x + width / 2, y - 14, "FECHA DE")
                    pdf.drawCentredString(x + width / 2, y - 25, "NACIMIENTO")
                else:
                    font_size = _fit_text(pdf, header, width - 5, bold_font, 9)
                    pdf.setFont(bold_font, font_size)
                    pdf.drawCentredString(x + width / 2, y - 21, header)
                x += width
            y -= header_height

            for row in page_rows:
                x = left
                for width, column in zip(widths, headers):
                    pdf.rect(x, y - row_height, width, row_height, fill=0, stroke=1)
                    value = row[column]
                    pdf.setFillColor(colors.black)
                    font_size = _fit_text(
                        pdf,
                        value,
                        width - 5,
                        regular_font,
                        min(9.5, max(6.0, row_height * 0.55)),
                    )
                    pdf.setFont(regular_font, font_size)
                    text_y = y - row_height / 2 - 2
                    pdf.drawCentredString(x + width / 2, text_y, _text(value))
                    x += width
                y -= row_height

            pdf.setFont(regular_font, 7)
            pdf.drawRightString(left + sum(widths), footer_margin - 5, f"Página {page_number} de {len(pages)}")
            pdf.showPage()
        pdf.save()

        buffer.seek(0)
        pdf_bytes = buffer.getvalue()
        if not pdf_bytes:
            raise ValueError("El PDF generado está vacío.")
        return pdf_bytes
    except Exception as error:
        raise RuntimeError(f"No se pudo generar el PDF: {error}") from error


def generate_pdf(
    records: pd.DataFrame,
    rows_per_page: int = DEFAULT_ROWS_PER_PAGE,
    logo_path: str | Path | None = None,
    authorized_document_exceptions: set[int] | frozenset[int] = frozenset(),
) -> bytes:
    """Generate a single PDF file containing all validated records."""
    return _build_pdf_bytes(
        final_records(records, authorized_document_exceptions),
        rows_per_page=rows_per_page,
        logo_path=logo_path,
    )


def generate_pdf_zip_by_unit(
    records: pd.DataFrame,
    rows_per_page: int = DEFAULT_ROWS_PER_PAGE,
    logo_path: str | Path | None = None,
    authorized_document_exceptions: set[int] | frozenset[int] = frozenset(),
    generated_on: date | None = None,
) -> bytes:
    """Generate a ZIP archive of separate PDFs, one per UNIDADES group."""
    try:
        frame = final_records(records, authorized_document_exceptions)
        stamp = (generated_on or date.today()).strftime("%d-%m-%Y")
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
            for unit, group in frame.groupby("UNIDADES", sort=False, dropna=False):
                unit_label = "NA" if is_missing(unit) else _sanitize_filename(unit)
                if not unit_label:
                    unit_label = "unidad"
                unit_frame = group.reset_index(drop=True)
                pdf_bytes = _build_pdf_bytes(
                    unit_frame,
                    rows_per_page=rows_per_page,
                    logo_path=logo_path,
                )
                archive.writestr(f"CERTIFICADOS_{unit_label}_{stamp}.pdf", pdf_bytes)

        buffer.seek(0)
        zip_bytes = buffer.getvalue()
        if not zip_bytes:
            raise ValueError("El ZIP generado está vacío.")
        return zip_bytes
    except Exception as error:
        raise RuntimeError(f"No se pudo generar el ZIP de PDFs por unidad: {error}") from error


def build_email_text(records: pd.DataFrame) -> str:
    """Build a default email body summarizing validation results.

    This helper is used in the Streamlit app to display the suggested notification
    text for the processed certificates and related observations.
    """
    validation = validate_records(records)
    active = validation["active"]
    missing_report = validation["missing_report"]
    return (
        "Asunto: Certificados procesados y novedades encontradas\n\n"
        "Buen día,\n\n"
        f"Se procesaron {len(active)} registros de ingreso. "
        f"Se identificaron {len(validation['duplicates'])} filas asociadas a documentos duplicados, "
        f"{len(validation['invalid_document'])} documentos con formato inválido y "
        f"{len(missing_report)} registros con uno o más campos obligatorios faltantes.\n\n"
        "Los casos fueron revisados por el responsable del proceso antes de generar el PDF final. "
        "Adjunto el listado organizado por UNIDADES y los reportes correspondientes.\n\n"
        "Quedo atenta a cualquier observación."
    )


def validation_summary(records: pd.DataFrame) -> dict[str, int | bool]:
    """Return a compact summary of validation metrics for the UI dashboard."""
    validation = validate_records(records)
    units = validation["active"]["UNIDADES"].map(lambda value: "NA" if is_missing(value) else value)
    problem_indexes = set(validation["invalid_document"].index)
    problem_indexes.update(validation["duplicates"].index)
    problem_indexes.update(validation["missing_report"].index)
    return {
        "registros_activos": len(validation["active"]),
        "registros_validos": len(validation["active"]) - len(problem_indexes),
        "documentos_faltantes": len(validation["missing_document"]),
        "documentos_invalidos": len(validation["invalid_document"]),
        "filas_duplicadas": len(validation["duplicates"]),
        "campos_informativos": len(validation["missing_report"]),
        "unidades": units.nunique(dropna=False),
        "blocking": bool(validation["blocking"]),
    }
