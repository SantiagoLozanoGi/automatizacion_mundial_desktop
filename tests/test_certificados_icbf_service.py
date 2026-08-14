from __future__ import annotations

from io import BytesIO
from datetime import date
import zipfile

import pandas as pd
from openpyxl import load_workbook
from pypdf import PdfReader

from workflows.certificados_icbf.service import CertificadosIcbfService
from app.file_io import save_bytes_to_file
from reportlab.pdfgen import canvas


def build_source() -> BytesIO:
    frame = pd.DataFrame(
        [
            ["Ana", None, "Díaz", None, "123", "01/01/2010", "Bogotá", "INGRESO"],
            ["Luis", "Alberto", "Pérez", None, "456", "02/02/2011", "Cali", "IN"],
        ],
        columns=[
            "Primer Nombre",
            "Segundo Nombre",
            "Primer Apellido",
            "Segundo Apellido",
            "Número de Identificación",
            "Fecha de Nacimiento",
            "UNIDADES",
            "Tipo de Novedad",
        ],
    )
    output = BytesIO()
    frame.to_excel(output, index=False)
    output.seek(0)
    return output


def test_service_processes_and_summarizes_excel() -> None:
    service = CertificadosIcbfService()

    records, stats = service.read_and_clean_excel(build_source())
    summary = service.validation_summary(records)

    assert stats == {"recibidos": 2, "ingresos": 2, "excluidos": 0}
    assert summary["registros_activos"] == 2
    assert summary["registros_validos"] == 2
    assert summary["blocking"] is False
    assert len(service.final_records(records)) == 2


def test_service_generates_pdf_and_zip_with_configured_logo() -> None:
    service = CertificadosIcbfService()
    records, _ = service.read_and_clean_excel(build_source())

    pdf_bytes = service.generate_pdf(records)
    zip_bytes = service.generate_pdf_zip_by_unit(records)

    assert pdf_bytes.startswith(b"%PDF")
    assert len(PdfReader(BytesIO(pdf_bytes)).pages) == 1
    with zipfile.ZipFile(BytesIO(zip_bytes)) as archive:
        assert sorted(archive.namelist()) == ["Bogotá.pdf", "Cali.pdf"]
        assert all(archive.read(name).startswith(b"%PDF") for name in archive.namelist())


def test_service_exposes_reports_and_email_text() -> None:
    service = CertificadosIcbfService()
    records, _ = service.read_and_clean_excel(build_source())
    validation = service.validate_records(records)

    excel_bytes = service.dataframe_to_excel_bytes(
        {"Duplicados": validation["duplicates"], "Faltantes": validation["missing_report"]}
    )
    email_text = service.build_email_text(records)

    assert excel_bytes.startswith(b"PK")
    assert "Se procesaron 2 registros" in email_text


def test_service_updates_include_without_mutating_original() -> None:
    service = CertificadosIcbfService()
    records, _ = service.read_and_clean_excel(build_source())

    updated = service.set_included(records, 0, False)

    assert bool(records.loc[0, "INCLUIR"]) is True
    assert bool(updated.loc[0, "INCLUIR"]) is False
    assert service.validation_summary(updated)["registros_activos"] == 1


def test_service_reports_row_anomalies_and_ready_state() -> None:
    service = CertificadosIcbfService()
    records, _ = service.read_and_clean_excel(build_source())
    records.loc[1, "DOCUMENTO"] = records.loc[0, "DOCUMENTO"]

    review = service.review_records(records)

    assert review["ready"] is False
    assert review["status"] == "Requiere revisión"
    assert review["summary"]["filas_duplicadas"] == 2
    assert all("duplicates" in review["rows"][row]["categories"] for row in (0, 1))
    assert "Documento duplicado" in review["rows"][0]["anomalies"][0]

    updated = service.set_included(records, 1, False)
    ready_review = service.review_records(updated)
    assert ready_review["ready"] is True
    assert ready_review["summary"]["registros_activos"] == 1
    assert ready_review["summary"]["no_seleccionados"] == 1
    assert ready_review["rows"][1]["status"] == "No incluido"


def test_service_reports_invalid_and_missing_fields() -> None:
    service = CertificadosIcbfService()
    records, _ = service.read_and_clean_excel(build_source())
    records.loc[0, "DOCUMENTO"] = "ABC"
    records.loc[0, "UNIDADES"] = ""

    row_review = service.review_records(records)["rows"][0]

    assert row_review["categories"] == {"included", "invalid", "missing"}
    assert any("Documento inválido" in item for item in row_review["anomalies"])
    assert any("UNIDADES" in item for item in row_review["anomalies"])


def test_review_session_reuses_validation_and_revalidates_business_edits(monkeypatch) -> None:
    import workflows.certificados_icbf.service as service_module

    workflow_service = CertificadosIcbfService()
    records, _ = workflow_service.read_and_clean_excel(build_source())
    real_validate = service_module.validate_records
    calls = 0

    def counted_validate(frame):
        nonlocal calls
        calls += 1
        return real_validate(frame)

    monkeypatch.setattr(service_module, "validate_records", counted_validate)
    session = workflow_service.create_review_session(records)
    initial_anomalies = session.snapshot()["rows"][0]["anomalies"]
    session.set_included(0, False)
    session.set_included(0, True)

    assert calls == 1
    assert session.snapshot()["rows"][0]["anomalies"] == initial_anomalies

    edited = session.records
    edited.loc[0, "DOCUMENTO"] = "INVALIDO"
    refreshed = session.revalidate(edited)
    assert calls == 2
    assert "invalid" in refreshed["rows"][0]["categories"]


def test_service_preserves_anomaly_reports_independently_of_selection() -> None:
    workflow_service = CertificadosIcbfService()
    records, _ = workflow_service.read_and_clean_excel(build_source())
    records.loc[1, "DOCUMENTO"] = records.loc[0, "DOCUMENTO"]
    records.loc[0, "UNIDADES"] = ""

    duplicates = workflow_service.generate_duplicates_report(records)
    missing = workflow_service.generate_missing_fields_report(records)
    availability = workflow_service.output_availability(records)

    assert duplicates.startswith(b"PK")
    assert missing.startswith(b"PK")
    assert availability["duplicates"] == 2
    assert availability["missing"] == 1
    assert availability["ready"] is False
    duplicate_book = load_workbook(BytesIO(duplicates), read_only=True)
    missing_book = load_workbook(BytesIO(missing), read_only=True)
    assert duplicate_book.sheetnames == ["Duplicados"]
    assert duplicate_book["Duplicados"].max_row == 3
    assert missing_book.sheetnames == ["Campos faltantes"]
    assert missing_book["Campos faltantes"].max_row == 2

    records.loc[1, "INCLUIR"] = False
    preserved = workflow_service.generate_duplicates_report(records)
    preserved_book = load_workbook(BytesIO(preserved), read_only=True)
    assert workflow_service.output_availability(records)["duplicates"] == 2
    assert preserved_book["Duplicados"].max_row == 3


def test_missing_report_keeps_excluded_anomaly_but_ignores_valid_excluded_row() -> None:
    workflow_service = CertificadosIcbfService()
    records, _ = workflow_service.read_and_clean_excel(build_source())
    records.loc[1, "UNIDADES"] = ""
    session = workflow_service.create_review_session(records)

    included_report = workflow_service.generate_missing_fields_report(session)
    session.set_included(1, False)
    excluded_report = workflow_service.generate_missing_fields_report(session)
    session.set_included(0, False)
    valid_excluded_report = workflow_service.generate_missing_fields_report(session)

    for content in (included_report, excluded_report, valid_excluded_report):
        workbook = load_workbook(BytesIO(content), read_only=True)
        sheet = workbook["Campos faltantes"]
        assert sheet.max_row == 2
        source_rows = [row[8] for row in sheet.iter_rows(values_only=True)]
        assert int(records.loc[1, "_FILA_ORIGEN"]) in source_rows
        assert int(records.loc[0, "_FILA_ORIGEN"]) not in source_rows


def test_service_output_names_and_file_writer(tmp_path) -> None:
    workflow_service = CertificadosIcbfService()
    filename = workflow_service.suggested_filename(
        "entrada real.xlsx", "pdf", generated_on=date(2026, 8, 13)
    )
    destination = save_bytes_to_file(tmp_path / filename, b"%PDF-test")

    assert filename == "CERTIFICADO-entrada real-20260813.pdf"
    assert destination.read_bytes() == b"%PDF-test"


def test_service_blocks_certificate_outputs_until_records_are_ready() -> None:
    workflow_service = CertificadosIcbfService()
    records, _ = workflow_service.read_and_clean_excel(build_source())
    records.loc[0, "DOCUMENTO"] = "INVALIDO"

    try:
        workflow_service.generate_pdf(records)
    except ValueError as error:
        assert "requieren revisión" in str(error)
    else:
        raise AssertionError("El service no debía generar un PDF bloqueado.")


def test_review_maps_missing_fields_to_original_source_row() -> None:
    workflow_service = CertificadosIcbfService()
    records, _ = workflow_service.read_and_clean_excel(build_source())
    records.loc[1, "UNIDADES"] = ""

    review = workflow_service.review_records(records)

    assert "missing" not in review["rows"][0]["categories"]
    assert "missing" in review["rows"][1]["categories"]
    updated = workflow_service.set_included(records, 1, False)
    assert workflow_service.output_availability(updated)["ready"] is True


def test_certificate_outputs_respect_include_selection() -> None:
    workflow_service = CertificadosIcbfService()
    records, _ = workflow_service.read_and_clean_excel(build_source())
    selected = workflow_service.set_included(records, 1, False)

    pdf = workflow_service.generate_pdf(selected)
    archive = workflow_service.generate_pdf_zip_by_unit(selected)
    pdf_text = "\n".join(page.extract_text() or "" for page in PdfReader(BytesIO(pdf)).pages)

    assert "Ana" in pdf_text
    assert "Luis" not in pdf_text
    with zipfile.ZipFile(BytesIO(archive)) as zipped:
        assert zipped.namelist() == ["Bogotá.pdf"]


def test_service_default_pdf_fits_seventy_rows_and_places_logo_left(monkeypatch) -> None:
    workflow_service = CertificadosIcbfService()
    records = pd.DataFrame(
        [
            {
                "INCLUIR": True,
                "PRIMER NOMBRE": f"Nombre {index}",
                "SEGUNDO NOMBRE": "NA",
                "PRIMER APELLIDO": "Apellido",
                "SEGUNDO APELLIDO": "NA",
                "DOCUMENTO": str(index + 1).zfill(10),
                "FECHA DE NACIMIENTO": "01/01/2010",
                "UNIDADES": "UNIDAD ÚNICA",
                "_FILA_ORIGEN": index + 2,
            }
            for index in range(70)
        ]
    )
    positions = []
    original_draw_image = canvas.Canvas.drawImage

    def capture_draw_image(pdf_canvas, image, x, y, *args, **kwargs):
        positions.append((x, y, kwargs["width"], kwargs["height"]))
        return original_draw_image(pdf_canvas, image, x, y, *args, **kwargs)

    monkeypatch.setattr(canvas.Canvas, "drawImage", capture_draw_image)
    pdf = workflow_service.generate_pdf(records)

    assert len(PdfReader(BytesIO(pdf)).pages) == 1
    assert positions and positions[0][0] == 25
    assert positions[0][2] / positions[0][3] > 1
